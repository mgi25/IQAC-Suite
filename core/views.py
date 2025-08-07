import os
import shutil
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponseForbidden, JsonResponse, HttpResponseRedirect
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.forms import inlineformset_factory
from django import forms
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import json
import logging
logger = logging.getLogger(__name__)
from .forms import RoleAssignmentForm, RegistrationForm
from .models import (
    Profile,
    RoleAssignment,
    Organization,
    OrganizationType,
    Report,
    Class,
    OrganizationRole,
)
from emt.models import EventProposal, Student
from django.views.decorators.http import require_GET, require_POST
from .models import ApprovalFlowTemplate, ApprovalFlowConfig
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────
def superuser_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if not (request.user.is_authenticated and request.user.is_superuser):
            return HttpResponseForbidden("You are not authorized to access this page.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def _superuser_check(user):
    if not user.is_superuser:
        raise PermissionDenied("You are not authorized to access this page.")
    return True


# Reuse the ModelForm defined in core.forms so it can be imported from here for tests


class RoleAssignmentFormSet(forms.BaseInlineFormSet):
    """Validate duplicate role assignments on the formset level."""

    def clean(self):
        super().clean()
        seen = set()
        for form in self.forms:
            if form.cleaned_data.get("DELETE"):
                continue
            role = form.cleaned_data.get("role")
            org = form.cleaned_data.get("organization")
            key = (role, org)
            if key in seen:
                raise forms.ValidationError(
                    "Duplicate role assignment for the same organization is not allowed."
                )
            seen.add(key)

# ─────────────────────────────────────────────────────────────
#  Auth Views
# ─────────────────────────────────────────────────────────────
def login_view(request):
    return render(request, "core/login.html")
    if user is not None:
        login(request, user)
        
        # Log the successful login
        logger.info(f"User '{user.username}' logged in successfully.")
        
        return redirect('dashboard')

def login_page(request):
    return render(request, "login.html")

def logout_view(request):
    logout(request)
    return redirect("login")

def custom_logout(request):
    logout(request)
    google_logout_url = "https://accounts.google.com/Logout"
    return redirect(f"{google_logout_url}?continue=https://appengine.google.com/_ah/logout?continue=http://127.0.0.1:8000/accounts/login/")


@login_required
def registration_form(request):
    """Collect registration number and role assignments for a user."""
    user = request.user
    domain = user.email.split("@")[-1].lower() if user.email else ""
    is_student = domain.endswith("christuniversity.in")

    student = None
    if is_student:
        student, _ = Student.objects.get_or_create(user=user)
        if student.registration_number:
            return redirect("dashboard")

    form_kwargs = {"include_regno": is_student}

    if request.method == "POST":
        form = RegistrationForm(request.POST, **form_kwargs)
        if form.is_valid():
            if is_student:
                student.registration_number = form.cleaned_data["registration_number"]
                student.save()
            for item in form.cleaned_data["assignments"]:
                RoleAssignment.objects.get_or_create(
                    user=user,
                    organization_id=item["organization"],
                    role_id=item["role"],
                )
            return redirect("dashboard")
    else:
        initial = {}
        if is_student:
            initial["registration_number"] = student.registration_number
        form = RegistrationForm(initial=initial, **form_kwargs)

    return render(request, "core/Registration_form.html", {"form": form, "is_student": is_student})


@require_GET
def api_organizations(request):
    """Return active organizations for typeahead."""
    q = request.GET.get("q", "").strip()
    orgs = Organization.objects.filter(is_active=True)
    if q:
        orgs = orgs.filter(name__icontains=q)
    data = [{"id": o.id, "text": o.name} for o in orgs.order_by("name")[:20]]
    return JsonResponse({"organizations": data})


@require_GET
def api_roles(request):
    """Return roles filtered by organization for typeahead."""
    org_id = request.GET.get("organization")
    q = request.GET.get("q", "").strip()
    roles = OrganizationRole.objects.filter(is_active=True)
    if org_id:
        roles = roles.filter(organization_id=org_id)
    if q:
        roles = roles.filter(name__icontains=q)
    data = [{"id": r.id, "text": r.name} for r in roles.order_by("name")[:20]]
    return JsonResponse({"roles": data})

# ─────────────────────────────────────────────────────────────
#  Dashboard
# ─────────────────────────────────────────────────────────────
@login_required
def dashboard(request):
    # Role-based landing logic
    user = request.user
    roles = RoleAssignment.objects.filter(user=user).select_related('role', 'organization')
    role_names = [ra.role.name for ra in roles]

    # Superuser/admin: redirect to admin dashboard
    if user.is_superuser:
        return redirect('admin_dashboard')

    # Student: show student dashboard (core/dashboard.html)
    if 'student' in [r.lower() for r in role_names]:
        dashboard_template = "core/dashboard.html"
    # Faculty: show faculty dashboard (core/dashboard.html)
    elif 'faculty' in [r.lower() for r in role_names]:
        dashboard_template = "core/dashboard.html"
    # Default: show dashboard
    else:
        dashboard_template = "core/dashboard.html"

    # Get events data
    finalized_events = EventProposal.objects.filter(status='finalized').distinct()
    
    # For students, show events they are participating in or have submitted
    # For faculty, show events they are involved with
    if 'student' in [r.lower() for r in role_names]:
        my_events = EventProposal.objects.filter(
            Q(submitted_by=user) | Q(status='finalized')
        ).distinct()
    else:
        my_events = finalized_events.filter(Q(submitted_by=user) | Q(faculty_incharges=user)).distinct()
    
    other_events = finalized_events.exclude(Q(submitted_by=user) | Q(faculty_incharges=user)).distinct()
    upcoming_events_count = finalized_events.filter(event_datetime__gte=timezone.now()).count()
    organized_events_count = EventProposal.objects.filter(submitted_by=user).count()
    week_start = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
    week_end = week_start + timezone.timedelta(days=6)
    this_week_events = finalized_events.filter(event_datetime__date__gte=week_start, event_datetime__date__lte=week_end).count()
    students_participated = finalized_events.aggregate(
        total=Sum('fest_fee_participants') + Sum('conf_fee_participants')
    )['total'] or 0
    my_students = Student.objects.filter(mentor=user)
    my_classes = Class.objects.filter(teacher=user)
    
    # Get user's recent proposals for notifications
    user_proposals = EventProposal.objects.filter(submitted_by=user).order_by('-updated_at')[:5]
    
    # Prepare calendar events data for JavaScript
    all_events = finalized_events.filter(event_datetime__isnull=False)
    calendar_events = []
    for event in all_events:
        calendar_events.append({
            'id': event.id,
            'title': event.event_title,
            'date': event.event_datetime.strftime('%Y-%m-%d'),
            'datetime': event.event_datetime.strftime('%Y-%m-%d %H:%M'),
            'venue': event.venue,
            'organization': event.organization.name if event.organization else '',
            'submitted_by': event.submitted_by.get_full_name() or event.submitted_by.username,
            'participants': event.fest_fee_participants or event.conf_fee_participants or 0,
            'is_my_event': user in [event.submitted_by] + list(event.faculty_incharges.all())
        })
    
    context = {
        "my_events": my_events,
        "other_events": other_events,
        "upcoming_events_count": upcoming_events_count,
        "organized_events_count": organized_events_count,
        "this_week_events": this_week_events,
        "students_participated": students_participated,
        "my_students": my_students,
        "my_classes": my_classes,
        "role_names": role_names,
        "user": user,
        "user_proposals": user_proposals,
        "calendar_events": json.dumps(calendar_events),
    }
    return render(request, dashboard_template, context)

@login_required
def propose_event(request):
    """
    This is the *old* quick proposal form. If you want to keep it, now use Organization instead of Department.
    """
    if request.method == "POST":
        org_type_name = request.POST.get("organization_type", "").strip()
        org_name = request.POST.get("organization", "").strip()
        organization = None
        if org_type_name and org_name:
            # Validate org_type_name exists to prevent typos creating unwanted entries
            try:
                org_type_obj = OrganizationType.objects.get(name=org_type_name)
            except OrganizationType.DoesNotExist:
                messages.error(request, f"Organization type '{org_type_name}' does not exist.")
                return redirect('propose_event')
            organization, _ = Organization.objects.get_or_create(name=org_name, org_type=org_type_obj)

        title = request.POST.get("title", "").strip()
        desc  = request.POST.get("description", "").strip()

        roles = [ra.role.name for ra in request.user.role_assignments.all()]
        user_type = ", ".join(roles) or getattr(request.user.profile, "role", "")

        EventProposal.objects.create(
            submitted_by=request.user,
            organization=organization,
            user_type=user_type,
            title=title,
            description=desc,
        )
        return redirect("dashboard")

    org_types = OrganizationType.objects.all().order_by('name')
    return render(request, "core/event_proposal.html", {"organization_types": org_types})

@login_required
def proposal_status(request, pk):
    proposal = get_object_or_404(EventProposal, pk=pk, submitted_by=request.user)
    steps = [
        {"key": "draft",     "label": "Draft"},
        {"key": "submitted", "label": "Submitted"},
        {"key": "under_review", "label": "Under Review"},
        {"key": "approved",  "label": "Approved"},
        {"key": "rejected",  "label": "Rejected"},
        {"key": "returned",  "label": "Returned for Revision"},
    ]
    return render(request, "core/proposal_status.html", {"proposal": proposal, "steps": steps})

# ─────────────────────────────────────────────────────────────
#  Admin Dashboard and User Management
# ─────────────────────────────────────────────────────────────
@login_required
def admin_dashboard(request):
    """
    Render the admin dashboard with dynamic analytics from the backend.
    """
    from django.contrib.auth.models import User
    from django.db.models import Q
    from datetime import timedelta

    # Calculate role statistics (manual count for accuracy)
    all_assignments = RoleAssignment.objects.select_related('role', 'user').all()
    counted_users = {'faculty': set(), 'student': set(), 'hod': set()}
    for assignment in all_assignments:
        role_name = assignment.role.name.lower()
        user_id = assignment.user.id
        if 'faculty' in role_name:
            counted_users['faculty'].add(user_id)
        elif 'student' in role_name:
            counted_users['student'].add(user_id)
        elif 'hod' in role_name or 'head' in role_name:
            counted_users['hod'].add(user_id)
    stats = {
        'students': len(counted_users['student']),
        'faculties': len(counted_users['faculty']),
        'hods': len(counted_users['hod']),
        'centers': Organization.objects.filter(is_active=True).count(),
        'departments': Organization.objects.filter(org_type__name__icontains='department', is_active=True).count(),
        'clubs': Organization.objects.filter(org_type__name__icontains='club', is_active=True).count(),
        'total_proposals': EventProposal.objects.count(),
        'pending_proposals': EventProposal.objects.filter(status__in=['submitted', 'under_review']).count(),
        'approved_proposals': EventProposal.objects.filter(status='approved').count(),
        'rejected_proposals': EventProposal.objects.filter(status='rejected').count(),
        'total_users': User.objects.count(),
        'active_users': User.objects.filter(is_active=True).count(),
        'new_users_this_week': User.objects.filter(date_joined__gte=timezone.now() - timedelta(days=7)).count(),
        'total_reports': Report.objects.count(),
        'database_status': 'Operational',
        'email_status': 'Active',
        'storage_status': '45% Used',
        'last_backup': timezone.now().strftime("%b %d, %Y"),
    }
    # Recent activity feed (proposals and reports)
    recent_activities = []
    recent_proposals = EventProposal.objects.select_related('submitted_by').order_by('-created_at')[:5]
    for proposal in recent_proposals:
        recent_activities.append({
            'type': 'proposal',
            'description': f"New event proposal: {getattr(proposal, 'event_title', getattr(proposal, 'title', 'Untitled Event'))}",
            'user': proposal.submitted_by.get_full_name() if proposal.submitted_by else '',
            'timestamp': proposal.created_at,
            'status': proposal.status
        })
    recent_reports = Report.objects.select_related('submitted_by').order_by('-created_at')[:3]
    for report in recent_reports:
        recent_activities.append({
            'type': 'report',
            'description': f"Report submitted: {report.title}",
            'user': report.submitted_by.get_full_name() if report.submitted_by else 'System',
            'timestamp': report.created_at,
            'status': getattr(report, 'status', '')
        })
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:8]
    context = {
        'stats': stats,
        'recent_activities': recent_activities,
    }
    return render(request, 'core/admin_dashboard.html', context)
@user_passes_test(lambda u: u.is_superuser)
def admin_user_panel(request):
    return render(request, "core/admin_user_panel.html")


# ===================================================================
# SINGLE-PAGE ROLE MANAGEMENT VIEW
# ===================================================================

@user_passes_test(lambda u: u.is_superuser)
def admin_role_management(request, organization_id=None):
    """Fixed role management with deduplication"""
    
    if organization_id:
        # Direct access to one organization's roles
        org = get_object_or_404(Organization, id=organization_id)
        roles = OrganizationRole.objects.filter(organization=org).order_by('name')
        assignments = (
            RoleAssignment.objects.filter(organization=org)
            .select_related("user", "role")
            .order_by("user__username")
        )

        context = {
            'selected_organization': org,
            'roles': roles,
            'role_assignments': assignments,
            'step': 'single_org_roles'
        }
        return render(request, "core/admin_role_management.html", context)
    
    else:
        org_type_id = request.GET.get('org_type_id')
        
        if org_type_id:
            # Show UNIQUE roles from this org type (no duplicates)
            org_type = get_object_or_404(OrganizationType, id=org_type_id)
            
            # Get all organizations of this type for the add form
            organizations = Organization.objects.filter(
                org_type=org_type, 
                is_active=True
            ).order_by('name')
            
            # Get UNIQUE role names (deduplicated) from all organizations of this type
            unique_role_names = OrganizationRole.objects.filter(
                organization__org_type=org_type,
                organization__is_active=True
            ).values_list('name', flat=True).distinct().order_by('name')
            
            # For each unique role name, get one representative role object
            unique_roles = []
            for role_name in unique_role_names:
                # Get the first role with this name from this org type
                role = OrganizationRole.objects.filter(
                    organization__org_type=org_type,
                    organization__is_active=True,
                    name=role_name
                ).select_related('organization').first()
                if role:
                    unique_roles.append(role)
            
            context = {
                'selected_org_type': org_type,
                'roles': unique_roles,  # Now deduplicated
                'organizations': organizations,
                'step': 'org_type_roles'
            }
            return render(request, "core/admin_role_management.html", context)
        
        else:
            # Show organization types
            org_types = OrganizationType.objects.all().order_by('name')
            
            context = {
                'org_types': org_types,
                'step': 'org_types'
            }
            return render(request, "core/admin_role_management.html", context)


@user_passes_test(lambda u: u.is_superuser)
@require_POST
def add_org_role(request, organization_id):
    """Add role to ALL organizations of the same type"""
    org = get_object_or_404(Organization, id=organization_id)
    name = request.POST.get("name", "").strip()
    description = request.POST.get("description", "").strip()

    if not name:
        messages.error(request, "Role name is required.")
        return redirect(f"{reverse('admin_role_management')}?org_type_id={org.org_type.id}")

    # Check if this role already exists in ANY organization of this type
    existing_role = OrganizationRole.objects.filter(
        organization__org_type=org.org_type,
        name__iexact=name
    ).first()
    
    if existing_role:
        messages.warning(request, f"Role '{name}' already exists for {org.org_type.name}s.")
        return redirect(f"{reverse('admin_role_management')}?org_type_id={org.org_type.id}")

    # Add this role to ALL organizations of this type
    all_orgs_of_type = Organization.objects.filter(
        org_type=org.org_type,
        is_active=True
    )
    
    for target_org in all_orgs_of_type:
        if not OrganizationRole.objects.filter(
            organization=target_org,
            name__iexact=name
        ).exists():
            OrganizationRole.objects.create(
                organization=target_org,
                name=name,
                is_active=True
                # Remove description if not in model
            )
    return redirect("admin_role_management") + f"?org_type_id={org.org_type.id}"


@user_passes_test(lambda u: u.is_superuser)
@require_POST
def add_role(request):
    """Add a role by organization or organization type (used by tests)."""
    name = request.POST.get("name", "").strip()
    org_id = request.POST.get("org_id")
    org_type_id = request.POST.get("org_type_id")

    if not name:
        return HttpResponseRedirect(reverse("admin_role_management"))

    if org_id:
        orgs = Organization.objects.filter(id=org_id, is_active=True)
        redirect_url = reverse("admin_role_management_org", args=[org_id])
    elif org_type_id:
        orgs = Organization.objects.filter(org_type_id=org_type_id, is_active=True)
        redirect_url = reverse("admin_role_management") + f"?org_type_id={org_type_id}"
    else:
        orgs = []
        redirect_url = reverse("admin_role_management")

    for org in orgs:
        if not OrganizationRole.objects.filter(organization=org, name__iexact=name).exists():
            OrganizationRole.objects.create(organization=org, name=name, is_active=True)

    return HttpResponseRedirect(redirect_url)


@user_passes_test(lambda u: u.is_superuser)
@require_POST
def update_org_role(request, role_id):
    """Update this role for ALL organizations of the same type"""
    role = get_object_or_404(OrganizationRole, id=role_id)
    org_type = role.organization.org_type
    old_name = role.name
    
    new_name = request.POST.get("name", "").strip()
    new_description = request.POST.get("description", "").strip()

    if not new_name:
        messages.error(request, "Role name is required.")
        return redirect("admin_role_management") + f"?org_type_id={org_type.id}"

    # Update all roles with the old name from all organizations of this type
    updated_count = OrganizationRole.objects.filter(
        organization__org_type=org_type,
        name__iexact=old_name
    ).update(
        name=new_name,
        description=new_description
    )
    
    messages.success(request, f"Role updated for {updated_count} {org_type.name}(s).")
    return redirect("admin_role_management") + f"?org_type_id={org_type.id}"


@user_passes_test(lambda u: u.is_superuser)
@require_POST
def toggle_org_role(request, role_id):
    """Handles activating or deactivating a role."""
    role = get_object_or_404(OrganizationRole, id=role_id)
    role.is_active = not role.is_active
    role.save()
    status = "activated" if role.is_active else "deactivated"
    messages.success(request, f"Role '{role.name}' has been {status}.")
    return redirect("admin_role_management_org", organization_id=role.organization.id)


@user_passes_test(lambda u: u.is_superuser)
@require_POST
def delete_org_role(request, role_id):
    """Delete this role from ALL organizations of the same type"""
    role = get_object_or_404(OrganizationRole, id=role_id)
    org_type = role.organization.org_type
    role_name = role.name
    
    # Delete all roles with this name from all organizations of this type
    deleted_count = OrganizationRole.objects.filter(
        organization__org_type=org_type,
        name__iexact=role_name
    ).delete()[0]
    
    messages.success(request, f"Role '{role_name}' deleted from {deleted_count} {org_type.name}(s).")
    return redirect("admin_role_management") + f"?org_type_id={org_type.id}"

# ===================================================================
# END OF SINGLE-PAGE ROLE MANAGEMENT
# ===================================================================

@user_passes_test(lambda u: u.is_superuser)
def admin_user_management(request):
    """
    Manages and displays a paginated list of users with filtering capabilities.
    """
    users_list = User.objects.select_related('profile').prefetch_related(
        'role_assignments__organization', 
        'role_assignments__role'
    ).order_by('-date_joined')

    q = request.GET.get('q', '').strip()
    role_id = request.GET.get('role')
    org_id = request.GET.get('organization')
    status = request.GET.get('status')

    if q:
        users_list = users_list.filter(
            Q(email__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(username__icontains=q)
        )
    
    if role_id:
        users_list = users_list.filter(role_assignments__role_id=role_id)
    
    if org_id:
        users_list = users_list.filter(role_assignments__organization_id=org_id)

    if status in ['active', 'inactive']:
        users_list = users_list.filter(is_active=(status == 'active'))

    users_list = users_list.distinct()

    paginator = Paginator(users_list, 25)
    page_number = request.GET.get('page')
    try:
        users = paginator.page(page_number)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    all_roles = OrganizationRole.objects.filter(is_active=True).order_by('name')
    all_organizations = Organization.objects.filter(is_active=True).order_by('name')

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    context = {
        'users': users,
        'all_roles': all_roles,
        'all_organizations': all_organizations,
        'current_filters': {
            'q': q,
            'role': int(role_id) if role_id and role_id.isdigit() else None,
            'organization': int(org_id) if org_id and org_id.isdigit() else None,
            'status': status,
        },
        'query_params': query_params.urlencode(),
    }
    
    return render(request, "core/admin_user_management.html", context)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_user_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)
    RoleFormSet = inlineformset_factory(
        User,
        RoleAssignment,
        form=RoleAssignmentForm,
        formset=RoleAssignmentFormSet,
        fields=("role", "organization"),
        extra=0,
        can_delete=True,
    )
    if request.method == "POST":
        formset = RoleFormSet(request.POST, instance=user)
        user.first_name = request.POST.get("first_name", "").strip()
        user.last_name  = request.POST.get("last_name", "").strip()
        user.email      = request.POST.get("email", "").strip()
        user.save()

        if formset.is_valid():
            formset.save()
            messages.success(request, "User roles updated successfully.")
            return redirect("admin_user_management")
        else:
            messages.error(request, "Please fix the errors below and try again.")
    else:
        formset = RoleFormSet(instance=user)

    org_ids = list(
        user.role_assignments.exclude(organization__isnull=True).values_list("organization_id", flat=True)
    )
    org_qs = Organization.objects.filter(Q(is_active=True) | Q(id__in=org_ids)).select_related("org_type")
    role_ids = list(user.role_assignments.values_list("role_id", flat=True))
    role_qs = OrganizationRole.objects.filter(Q(is_active=True) | Q(id__in=role_ids)).select_related("organization")

    organizations_json = {
        str(o.id): {"name": o.name, "org_type": str(o.org_type_id)} for o in org_qs
    }
    roles_json = {
        str(r.id): {"name": r.name, "organization": str(r.organization_id)} for r in role_qs
    }

    return render(
        request,
        "core/admin_user_edit.html",
        {
            "user_obj": user,
            "formset": formset,
            "organizations": org_qs,
            "organization_types": OrganizationType.objects.filter(),
            "organizations_json": json.dumps(organizations_json),
            "roles_json": json.dumps(roles_json),
        },
    )

@user_passes_test(lambda u: u.is_superuser)
def admin_event_proposals(request):
    proposals = EventProposal.objects.all().order_by("-created_at")
    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    if q:
        proposals = proposals.filter(
            Q(event_title__icontains=q) |
            Q(submitted_by__username__icontains=q) |
            Q(organization__name__icontains=q) |
            Q(organization__org_type__name__icontains=q)
        )
    if status:
        proposals = proposals.filter(status=status)
    return render(request, "core/admin_event_proposals.html", {"proposals": proposals})

@user_passes_test(lambda u: u.is_superuser)
def event_proposal_json(request, proposal_id):
    p = get_object_or_404(EventProposal, id=proposal_id)
    return JsonResponse({
        "title": p.title,
        "description": p.description,
        "organization": str(p.organization) if p.organization else None,
        "user_type": p.user_type,
        "status": p.status,
        "status_display": p.get_status_display(),
        "date_submitted": p.date_submitted.strftime("%Y-%m-%d %H:%M"),
        "submitted_by": p.submitted_by.get_full_name() or p.submitted_by.username,
    })

@user_passes_test(lambda u: u.is_superuser)
@require_POST
def event_proposal_action(request, proposal_id):
    p = get_object_or_404(EventProposal, id=proposal_id)
    data = json.loads(request.body)
    action = data.get("action")
    comment = data.get("comment", "")
    if action == "approved":
        p.status = "approved"
    elif action == "rejected":
        p.status = "rejected"
    elif action == "returned":
        p.status = "returned"
    else:
        return JsonResponse({"success": False, "error": "Invalid action"})
    if comment:
        p.admin_comment = comment
    p.save()
    return JsonResponse({"success": True})

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_reports(request):
    reports = Report.objects.select_related('submitted_by').all().order_by('-created_at')
    return render(request, 'core/admin_reports.html', {
        'reports': reports,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_reports_approve(request, report_id):
    report = get_object_or_404(Report, id=report_id)
    report.status = 'approved'
    report.save()
    messages.success(request, f"Report '{report.title}' approved.")
    return HttpResponseRedirect(reverse('admin_reports'))

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_reports_reject(request, report_id):
    report = get_object_or_404(Report, id=report_id)
    report.status = 'rejected'
    report.save()
    messages.warning(request, f"Report '{report.title}' rejected.")
    return HttpResponseRedirect(reverse('admin_reports'))

def cdl_dashboard(request):
    return render(request, 'emt/cdl_dashboard.html')

STATUSES = ['submitted', 'under_review', 'approved', 'rejected']

from emt.models import EventProposal

STATUSES = ['submitted', 'under_review', 'rejected', 'finalized']  # ensure this matches your status choices

def iqac_suite_dashboard(request):
    user_proposals = (
        EventProposal.objects
        .filter(submitted_by=request.user)
        .order_by("-created_at")
    )

    for proposal in user_proposals:
        proposal.statuses = STATUSES
        proposal.status_index = STATUSES.index(proposal.status) if proposal.status in STATUSES else -1
        proposal.progress_percent = ((proposal.status_index + 1) / len(STATUSES)) * 100 if proposal.status_index != -1 else 0
        proposal.current_label = proposal.status.replace('_', ' ').title()

    context = {
        "user_proposals": user_proposals,
        "statuses": STATUSES,
    }
    return render(request, "dashboard.html", context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_master_data(request):
    from transcript.models import AcademicYear
    import datetime
    import json

    selected_year_param = request.GET.get('year')
    current_year = datetime.datetime.now().year

    academic_years_from_db = AcademicYear.objects.all().order_by('-year')
    if not academic_years_from_db.exists():
        for year in range(current_year - 1, current_year + 3):
            AcademicYear.objects.create(year=f"{year}-{year + 1}")
        academic_years_from_db = AcademicYear.objects.all().order_by('-year')

    academic_years = [{'value': ay.year, 'display': ay.year} for ay in academic_years_from_db]

    if selected_year_param:
        selected_year = next((ay for ay in academic_years if ay['value'] == selected_year_param), None)
        if not selected_year:
            selected_year = academic_years[0] if academic_years else {'value': f"{current_year}-{current_year + 1}", 'display': f"{current_year}-{current_year + 1}"}
    else:
        selected_year = academic_years[0] if academic_years else {'value': f"{current_year}-{current_year + 1}", 'display': f"{current_year}-{current_year + 1}"}

    org_types = OrganizationType.objects.filter(is_active=True).order_by('name')
    orgs_by_type = {}
    orgs_by_type_json = {}

    for org_type in org_types:
        qs = Organization.objects.filter(org_type=org_type).order_by('name')
        orgs_by_type[org_type.name] = qs
        active_orgs = qs.filter(is_active=True)
        orgs_by_type_json[org_type.name.lower()] = [{'id': o.id, 'name': o.name} for o in active_orgs]

    return render(request, "core/admin_master_data.html", {
        "org_types": org_types,
        "orgs_by_type": orgs_by_type,
        "academic_years": academic_years,
        "selected_year": selected_year,
        "orgs_by_type_json": json.dumps(orgs_by_type_json),
    })



@login_required
@user_passes_test(lambda u: u.is_superuser)
@csrf_exempt
def admin_master_data_edit(request, model_name, pk):
    MODEL_MAP = {
        "organization": Organization,
        "organization_type": OrganizationType,
    }
    Model = MODEL_MAP.get(model_name)
    if request.method == "POST" and Model:
        try:
            obj = get_object_or_404(Model, pk=pk)
            data = json.loads(request.body)
            name = data.get("name", "").strip()
            is_active = data.get("is_active", True)
            parent_id = data.get("parent")
            
            if not name:
                return JsonResponse({"success": False, "error": "Name is required"})
            
            obj.name = name
            if hasattr(obj, "is_active"):
                obj.is_active = is_active
            
            if model_name == "organization" and hasattr(obj, "parent"):
                if parent_id and parent_id.strip():
                    try:
                        parent_obj = Organization.objects.get(id=parent_id)
                        if obj.org_type.parent_type and parent_obj.org_type != obj.org_type.parent_type:
                            return JsonResponse({
                                "success": False, 
                                "error": f"Invalid parent type. Expected {obj.org_type.parent_type.name}, got {parent_obj.org_type.name}"
                            })
                        obj.parent = parent_obj
                    except Organization.DoesNotExist:
                        return JsonResponse({"success": False, "error": "Parent organization not found"})
                    except ValueError:
                        return JsonResponse({"success": False, "error": "Invalid parent ID"})
                else:
                    obj.parent = None
            
            obj.save()
            
            response_data = {
                "success": True,
                "name": obj.name,
                "is_active": getattr(obj, "is_active", True),
            }
            
            if hasattr(obj, "parent"):
                response_data["parent"] = obj.parent.name if obj.parent else None
            
            return JsonResponse(response_data)
            
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "Invalid JSON data"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})

@login_required
@user_passes_test(lambda u: u.is_superuser)
@csrf_exempt
def admin_master_data_add(request, model_name):
    MODEL_MAP = {
        "organization": Organization,
        "organization_type": OrganizationType,
    }
    Model = MODEL_MAP.get(model_name)
    if request.method == "POST" and Model:
        try:
            data = json.loads(request.body)
            name = data.get("name", "").strip()
            
            if not name:
                return JsonResponse({"success": False, "error": "Name is required"})
            
            if model_name == "organization":
                org_type_name = data.get("org_type", "").strip()
                parent_id = data.get("parent")
                
                if not org_type_name:
                    return JsonResponse({"success": False, "error": "Organization type required"})
                
                try:
                    org_type_obj = OrganizationType.objects.get(name__iexact=org_type_name)
                except OrganizationType.DoesNotExist:
                    return JsonResponse({"success": False, "error": f"Organization type '{org_type_name}' does not exist"})
                
                if Organization.objects.filter(name__iexact=name, org_type=org_type_obj).exists():
                    return JsonResponse({"success": False, "error": f"{org_type_obj.name} with name '{name}' already exists"})
                
                parent_obj = None
                if parent_id:
                    try:
                        parent_obj = Organization.objects.get(id=parent_id)
                    except Organization.DoesNotExist:
                        return JsonResponse({"success": False, "error": "Parent organization not found"})
                
                obj = Organization.objects.create(
                    name=name,
                    org_type=org_type_obj,
                    parent=parent_obj,
                    is_active=True
                )
                
            elif model_name == "organization_type":
                parent_id = data.get("parent")
                
                if OrganizationType.objects.filter(name__iexact=name).exists():
                    return JsonResponse({"success": False, "error": f"Organization type '{name}' already exists"})
                
                parent_obj = None
                if parent_id:
                    try:
                        parent_obj = OrganizationType.objects.get(id=parent_id)
                    except OrganizationType.DoesNotExist:
                        return JsonResponse({"success": False, "error": "Parent organization type not found"})
                
                obj = OrganizationType.objects.create(
                    name=name,
                    parent_type=parent_obj,
                    can_have_parent=bool(parent_obj),
                    is_active=True
                )
            else:
                obj, created = Model.objects.get_or_create(name=name, defaults={'is_active': True})
                if not created:
                    return JsonResponse({"success": False, "error": f"{model_name} already exists"})
            
            return JsonResponse({
                "success": True,
                "id": obj.id,
                "name": obj.name,
                "org_type": obj.org_type.name if model_name == "organization" else None,
                "parent": obj.parent.name if hasattr(obj, 'parent') and obj.parent else None,
                "is_active": getattr(obj, 'is_active', True)
            })
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "Invalid JSON data"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required
@user_passes_test(lambda u: u.is_superuser)
@csrf_exempt
def admin_master_data_delete(request, model_name, pk):
    MODEL_MAP = {
        "organization": Organization,
        "organization_type": OrganizationType,
    }
    Model = MODEL_MAP.get(model_name)
    if request.method == "POST" and Model:
        obj = get_object_or_404(Model, pk=pk)
        obj.delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "error": "Invalid request"})

@user_passes_test(lambda u: u.is_superuser)
def admin_proposal_detail(request, proposal_id):
    proposal = get_object_or_404(
        EventProposal.objects.select_related("organization").prefetch_related(
            "faculty_incharges", "speakers", "expense_details", "approval_steps"
        ),
        id=proposal_id,
    )
    logger.debug(
        "Loaded proposal %s '%s' (org=%s, faculty=%d, speakers=%d, expenses=%d, steps=%d)",
        proposal.id,
        proposal.event_title,
        proposal.organization,
        len(proposal.faculty_incharges.all()),
        len(proposal.speakers.all()),
        len(proposal.expense_details.all()),
        len(proposal.approval_steps.all()),
    )
    return render(request, "core/admin_proposal_detail.html", {"proposal": proposal})

@login_required
def proposal_detail(request, proposal_id):
    """Public view for event proposal details - accessible to all logged-in users"""
    proposal = get_object_or_404(
        EventProposal.objects.select_related("organization").prefetch_related(
            "faculty_incharges", "speakers", "expense_details", "approval_steps"
        ),
        id=proposal_id,
    )
    logger.debug(
        "Loaded public proposal %s '%s' (org=%s, faculty=%d, speakers=%d, expenses=%d, steps=%d)",
        proposal.id,
        proposal.event_title,
        proposal.organization,
        len(proposal.faculty_incharges.all()),
        len(proposal.speakers.all()),
        len(proposal.expense_details.all()),
        len(proposal.approval_steps.all()),
    )
    return render(request, "core/admin_proposal_detail.html", {"proposal": proposal})

@user_passes_test(lambda u: u.is_superuser)
def admin_settings_dashboard(request):
    return render(request, "core/admin_settings.html")

@user_passes_test(lambda u: u.is_superuser)
def master_data_dashboard(request):
    from transcript.models import AcademicYear
    from django.contrib.auth.models import User
    
    stats = {
        'organizations': Organization.objects.count(),
        'org_types': OrganizationType.objects.count(),
        'academic_years': AcademicYear.objects.count(),
        'active_users': User.objects.filter(is_active=True).count(),
    }
    
    recent_activities = [
        {
            'description': 'System initialized',
            'icon': 'info-circle',
            'created_at': timezone.now(),
        }
    ]
    
    return render(request, "core/master_data_dashboard.html", {
        'stats': stats,
        'recent_activities': recent_activities,
    })

@user_passes_test(lambda u: u.is_superuser)
def admin_approval_flow(request):
    """List all organizations grouped by type for approval flow management."""
    org_types = OrganizationType.objects.filter(is_active=True).order_by("name")
    orgs_by_type = {
        ot.name: Organization.objects.filter(org_type=ot, is_active=True).order_by("name")
        for ot in org_types
    }
    context = {
        "org_types": org_types,
        "orgs_by_type": orgs_by_type,
    }
    return render(request, "core/admin_approval_flow_list.html", context)


@user_passes_test(lambda u: u.is_superuser)
def admin_approval_flow_manage(request, org_id):
    """Display and edit approval flow for a single organization."""
    orgs = (
        Organization.objects.filter(is_active=True)
        .select_related("org_type")
        .order_by("org_type__name", "name")
    )
    org_types = OrganizationType.objects.all().order_by("name")
    selected_org = get_object_or_404(Organization, id=org_id)

    config, _ = ApprovalFlowConfig.objects.get_or_create(organization=selected_org)

    steps = (
        ApprovalFlowTemplate.objects.filter(organization=selected_org)
        .select_related("user")
        .order_by("step_order")
    )

    context = {
        "organizations": orgs,
        "org_types": org_types,
        "selected_org_id": org_id,
        "selected_org": selected_org,
        "existing_steps": steps,
        "require_faculty_incharge_first": config.require_faculty_incharge_first,
    }
    return render(request, "core/admin_approval_flow_manage.html", context)


@user_passes_test(lambda u: u.is_superuser)
def admin_approval_dashboard(request):
    """Intermediate page for approval related settings."""
    return render(request, "core/admin_approval_dashboard.html")




@login_required
@csrf_exempt
def set_academic_year(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            academic_year = data.get('academic_year')
            if academic_year:
                request.session['selected_academic_year'] = str(academic_year)
                return JsonResponse({'success': True})
        except (json.JSONDecodeError, ValueError):
            pass
    return JsonResponse({'success': False})

@login_required
@user_passes_test(lambda u: u.is_superuser)
@csrf_exempt
def add_academic_year(request):
    if request.method == 'POST':
        try:
            import re
            from transcript.models import AcademicYear
            data = json.loads(request.body)
            academic_year = data.get('academic_year')
            if not academic_year:
                return JsonResponse({'success': False, 'error': 'Academic year is required'})
            if not re.match(r'^\d{4}-\d{4}$', academic_year):
                return JsonResponse({'success': False, 'error': 'Invalid format. Use YYYY-YYYY (e.g., 2025-2026)'})
            if AcademicYear.objects.filter(year=academic_year).exists():
                return JsonResponse({'success': False, 'error': f'Academic year {academic_year} already exists'})
            AcademicYear.objects.create(year=academic_year)
            return JsonResponse({'success': True, 'message': f'Academic year {academic_year} added successfully'})
        except (json.JSONDecodeError, ValueError) as e:
            return JsonResponse({'success': False, 'error': 'Invalid request data'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Only POST method allowed'})

@user_passes_test(lambda u: u.is_superuser)
def admin_pso_po_management(request):
    import json
    from .models import Program, ProgramOutcome, ProgramSpecificOutcome
    
    org_types = OrganizationType.objects.filter(is_active=True).order_by('name')
    orgs_by_type = {}
    orgs_by_type_json = {}

    for org_type in org_types:
        qs = Organization.objects.filter(org_type=org_type).order_by('name')
        orgs_by_type[org_type.name] = qs
        active_orgs = qs.filter(is_active=True)
        orgs_by_type_json[org_type.name.lower()] = [{'id': o.id, 'name': o.name} for o in active_orgs]
    
    programs = Program.objects.all().order_by("name")
    context = {
        "org_types": org_types,
        "orgs_by_type": orgs_by_type,
        "orgs_by_type_json": json.dumps(orgs_by_type_json),
        "programs": programs,
    }
    return render(request, "core/admin_pso_po_management.html", context)

@require_GET
def get_approval_flow(request, org_id):
    steps = ApprovalFlowTemplate.objects.filter(organization_id=org_id).order_by('step_order')
    config = ApprovalFlowConfig.objects.filter(organization_id=org_id).first()
    data = [
        {
            'id': step.id,
            'step_order': step.step_order,
            'role_required': step.role_required,
            'role_display': step.get_role_required_display(),
            'user_id': step.user.id if step.user else None,
            'user_name': step.user.get_full_name() if step.user else '',
            'optional': step.optional,
        } for step in steps
    ]
    return JsonResponse({'success': True, 'steps': data,
                         'require_faculty_incharge_first': config.require_faculty_incharge_first if config else False})

@require_POST
@login_required
@user_passes_test(lambda u: _superuser_check(u))
def save_approval_flow(request, org_id):
    data = json.loads(request.body)
    steps = data.get('steps', [])
    require_first = data.get('require_faculty_incharge_first', False)
    org = Organization.objects.get(id=org_id)
    ApprovalFlowTemplate.objects.filter(organization=org).delete()
    for idx, step in enumerate(steps, 1):
        if not step.get('role_required'):
            continue
        ApprovalFlowTemplate.objects.create(
            organization=org,
            step_order=idx,
            role_required=step['role_required'],
            user_id=step.get('user_id'),
            optional=step.get('optional', False)
        )

    config, _ = ApprovalFlowConfig.objects.get_or_create(organization=org)
    config.require_faculty_incharge_first = require_first
    config.save()

    return JsonResponse({'success': True})
def api_approval_flow_steps(request, org_id):
    steps = list(
        ApprovalFlowTemplate.objects.filter(organization_id=org_id).order_by('step_order').values(
            'id', 'step_order', 'role_required', 'user_id', 'optional'
        )
    )
    return JsonResponse({'success': True, 'steps': steps})

@require_POST
@csrf_exempt
@user_passes_test(lambda u: u.is_superuser)
def delete_approval_flow(request, org_id):
    ApprovalFlowTemplate.objects.filter(organization_id=org_id).delete()
    return JsonResponse({'success': True})
@login_required
@user_passes_test(lambda u: u.is_superuser)
def search_users(request):
    q = request.GET.get("q", "").strip()
    role = request.GET.get("role", "").strip()
    org_id = request.GET.get("org_id", "").strip()
    org_type_id = request.GET.get("org_type_id", "").strip()

    users = User.objects.all()

    if q:
        users = users.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
        )

    # --- Filter by role ---
    if role:
        users = users.filter(role_assignments__role__name__iexact=role)

    # --- Filter by organization/department ---
    if org_id:
        users = users.filter(role_assignments__organization_id=org_id)

    users = users.distinct()[:10]
    data = [
        {"id": u.id, "name": u.get_full_name() or u.username, "email": u.email}
        for u in users
    ]
    if not data:
        logging.debug(
            "search_users returned no results for q=%r role=%r org_id=%r org_type_id=%r",
            q,
            role,
            org_id,
            org_type_id,
        )
    return JsonResponse({"success": True, "users": data})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def organization_users(request, org_id):
    """Return users for a given organization, optional search by name or role."""
    q = request.GET.get("q", "")
    role = request.GET.get("role", "")
    org_type_id = request.GET.get("org_type_id", "")

    assignments = RoleAssignment.objects.filter(organization_id=org_id)
    if not assignments.exists() and org_type_id:
        assignments = RoleAssignment.objects.filter(organization__org_type_id=org_type_id)
    if role:
        assignments = assignments.filter(role__name__iexact=role)
    if q:
        assignments = assignments.filter(
            Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(user__email__icontains=q)
        )

    assignments = assignments.select_related("user").order_by("user__first_name", "user__last_name")

    users_data = [
        {
            "id": a.user.id,
            "name": a.user.get_full_name() or a.user.username,
            "role": a.role.name,
        }
        for a in assignments
    ]
    return JsonResponse({"success": True, "users": users_data})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def api_org_type_organizations(request, org_type_id):
    """Return active organizations for a given organization type."""
    orgs = Organization.objects.filter(org_type_id=org_type_id, is_active=True).order_by("name")
    data = [{"id": o.id, "name": o.name} for o in orgs]
    return JsonResponse({"success": True, "organizations": data})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def api_org_type_roles(request, org_type_id):
    """Return distinct role names for a given organization type."""
    roles = (
        OrganizationRole.objects
        .filter(organization__org_type_id=org_type_id, is_active=True)
        .values("id", "name")
        .order_by("name")
    )
    # Deduplicate by name while preserving an ID for each
    seen = set()
    unique_roles = []
    for r in roles:
        if r["name"] not in seen:
            seen.add(r["name"])
            unique_roles.append({"id": r["id"], "name": r["name"]})
    return JsonResponse({"success": True, "roles": unique_roles})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def api_organization_roles(request, org_id):
    """Return role names for a specific organization."""
    roles = (
        OrganizationRole.objects
        .filter(organization_id=org_id, is_active=True)
        .values("id", "name")
        .order_by("name")
    )
    data = list(roles)
    return JsonResponse({"success": True, "roles": data})
# PSO/PO Management API endpoints
from django.views.decorators.http import require_GET, require_POST

@require_GET
@user_passes_test(lambda u: u.is_superuser)
def get_pso_po_data(request, org_type, org_id):
    """Get POs and PSOs for a specific organization"""
    try:
        from .models import Program, ProgramOutcome, ProgramSpecificOutcome
        
        org = Organization.objects.get(id=org_id)
        
        programs = Program.objects.filter(organization=org)
        
        pos = []
        psos = []
        
        if programs.exists():
            program = programs.first()
            pos = list(ProgramOutcome.objects.filter(program=program).values('id', 'description'))
            psos = list(ProgramSpecificOutcome.objects.filter(program=program).values('id', 'description'))
        
        return JsonResponse({
            'success': True,
            'pos': pos,
            'psos': psos,
            'organization': org.name
        })
    except Organization.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Organization not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
@csrf_exempt
@user_passes_test(lambda u: u.is_superuser)
def add_outcome(request, outcome_type):
    """Add a new PO or PSO"""
    try:
        from .models import Program, ProgramOutcome, ProgramSpecificOutcome
        
        data = json.loads(request.body)
        org_id = data.get('org_id')
        description = data.get('description')
        
        if not org_id or not description:
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        org = Organization.objects.get(id=org_id)
        
        program, created = Program.objects.get_or_create(
            organization=org,
            defaults={'name': f"{org.name} Program"}
        )
        
        if outcome_type == 'po':
            ProgramOutcome.objects.create(program=program, description=description)
        elif outcome_type == 'pso':
            ProgramSpecificOutcome.objects.create(program=program, description=description)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid outcome type'})
        
        return JsonResponse({'success': True})
        
    except Organization.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Organization not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_POST
@csrf_exempt
@user_passes_test(lambda u: u.is_superuser)
def delete_outcome(request, outcome_type, outcome_id):
    """Delete a PO or PSO"""
    try:
        from .models import ProgramOutcome, ProgramSpecificOutcome
        
        if outcome_type == 'po':
            outcome = ProgramOutcome.objects.get(id=outcome_id)
        elif outcome_type == 'pso':
            outcome = ProgramSpecificOutcome.objects.get(id=outcome_id)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid outcome type'})
        
        outcome.delete()
        return JsonResponse({'success': True})
        
    except (ProgramOutcome.DoesNotExist, ProgramSpecificOutcome.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Outcome not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from itertools import chain
from operator import attrgetter
from core.models import Report  # instead of SubmittedReport
from emt.models import EventReport

@login_required
def admin_reports_view(request):
    try:
        # Use Report instead of SubmittedReport here:
        submitted_reports = Report.objects.all()

        generated_reports = EventReport.objects.select_related('proposal').all()

        all_reports_list = list(chain(submitted_reports, generated_reports))

        all_reports_list.sort(key=attrgetter('created_at'), reverse=True)

        context = {'reports': all_reports_list}

        return render(request, 'core/admin_reports.html', context)

    except Exception as e:
        print(f"Error in admin_reports_view: {e}")
        return HttpResponse(f"An error occurred: {e}", status=500)

# ======================== API Endpoints & User Dashboard ========================

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET

@login_required
def api_auth_me(request):
    user = request.user
    role = 'faculty' if user.is_staff else 'student'
    initials = ''.join([x[0] for x in user.get_full_name().split()]) or user.username[:2].upper()
    return JsonResponse({
        'role': role,
        'name': user.get_full_name(),
        'subtitle': '',  # Add more info if needed
        'initials': initials,
    })

@login_required
def api_faculty_overview(request):
    stats = [
        # Build from your models as needed
    ]
    return JsonResponse(stats, safe=False)

@login_required
def user_dashboard(request):
    return render(request, 'core/user_dashboard.html')

@login_required
@require_GET
def api_global_search(request):
    """
    Global search API endpoint for the Central Command Center.
    Searches across Students, Event Proposals, Reports, and Users.
    """
    from django.db.models import Q
    from django.contrib.auth.models import User
    import json
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({
            'success': True,
            'results': {
                'students': [],
                'proposals': [],
                'reports': [],
                'users': []
            }
        })
    try:
        results = {'students': [], 'proposals': [], 'reports': [], 'users': []}
        try:
            from transcript.models import Student
            students = Student.objects.filter(
                Q(name__icontains=query) |
                Q(roll_no__icontains=query)
            ).select_related('school', 'course')[:5]
            results['students'] = [{
                'id': student.id,
                'name': student.name,
                'roll_no': student.roll_no,
                'school': student.school.name if student.school else 'N/A',
                'course': student.course.name if student.course else 'N/A',
                'url': f'/transcript/{student.roll_no}/'
            } for student in students]
        except ImportError:
            student_users = User.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(username__icontains=query)
            ).filter(
                profile__isnull=False
            )[:5]
            results['students'] = [{
                'id': user.id,
                'name': f"{user.first_name} {user.last_name}".strip() or user.username,
                'roll_no': user.username,
                'school': 'N/A',
                'course': 'N/A',
                'url': f'/core-admin/users/{user.id}/'
            } for user in student_users]
        try:
            from emt.models import EventProposal
            proposals = EventProposal.objects.filter(
                Q(event_title__icontains=query) |
                Q(submitted_by_first_name_icontains=query) |
                Q(submitted_by_last_name_icontains=query) |
                Q(organization_name_icontains=query) |
                Q(event_focus_type__icontains=query)
            ).select_related('submitted_by', 'organization').order_by('-event_datetime')[:5]
            if query.lower() == 'iqac':
                proposals = EventProposal.objects.filter(
                    Q(organization_name_icontains='iqac') |
                    Q(event_focus_type__icontains='iqac')
                ).select_related('submitted_by', 'organization').order_by('-event_datetime')[:5]
            results['proposals'] = [{
                'id': proposal.id,
                'title': proposal.event_title,
                'faculty': proposal.submitted_by.get_full_name() if proposal.submitted_by else 'N/A',
                'organization': proposal.organization.name if proposal.organization else 'N/A',
                'status': getattr(proposal, 'status', 'Unknown'),
                'date': proposal.event_datetime.strftime('%Y-%m-%d') if proposal.event_datetime else 'N/A',
                'url': f'/core-admin/event-proposals/{proposal.id}/'
            } for proposal in proposals]
        except (ImportError, AttributeError):
            results['proposals'] = []
        try:
            from core.models import Report
            reports = Report.objects.filter(
                Q(title__icontains=query) |
                Q(description__icontains=query) |
                Q(organization_name_icontains=query)
            ).order_by('-created_at')[:5]
            if query.lower() == 'iqac':
                reports = Report.objects.filter(
                    Q(report_type__iexact='iqac') |
                    Q(title__icontains='iqac') |
                    Q(organization_name_icontains='iqac')
                ).order_by('-created_at')[:5]
            results['reports'] = [{
                'id': report.id,
                'title': report.title,
                'type': getattr(report, 'report_type', 'Report'),
                'organization': report.organization.name if report.organization else 'N/A',
                'date': report.created_at.strftime('%Y-%m-%d'),
                'url': f'/core-admin/reports/{report.id}/'
            } for report in reports]
        except (ImportError, AttributeError):
            results['reports'] = []
        if request.user.is_superuser or hasattr(request.user, 'profile'):
            users = User.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(username__icontains=query) |
                Q(email__icontains=query)
            ).exclude(id=request.user.id)[:5]
            results['users'] = [{
                'id': user.id,
                'name': f"{user.first_name} {user.last_name}".strip() or user.username,
                'email': user.email,
                'role': getattr(user.profile, 'role', 'User') if hasattr(user, 'profile') else 'User',
                'last_login': user.last_login.strftime('%Y-%m-%d') if user.last_login else 'Never',
                'url': f'/core-admin/users/{user.id}/edit/'
            } for user in users]
        return JsonResponse({'success': True, 'results': results, 'query': query})
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'results': {
                'students': [],
                'proposals': [],
                'reports': [],
                'users': []
            }
        }, status=500)

@login_required
@require_GET
def admin_dashboard_api(request):
    """
    API endpoint for dashboard analytics - useful for real-time updates.
    """
    from django.contrib.auth.models import User
    stats = {
        'students': User.objects.filter(
            role_assignments__role__name__icontains='student'
        ).distinct().count(),
        'faculties': User.objects.filter(
            role_assignments__role__name__icontains='faculty'
        ).distinct().count(),
        'hods': User.objects.filter(
            role_assignments__role__name__icontains='hod'
        ).distinct().count(),
        'centers': Organization.objects.filter(is_active=True).count(),
        'total_users': User.objects.filter(is_active=True).count(),
        'total_proposals': EventProposal.objects.count(),
        'pending_proposals': EventProposal.objects.filter(
            status__in=['submitted', 'under_review']
        ).count(),
    }
    return JsonResponse({'success': True, 'stats': stats})
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db.models import Q, Count, Prefetch
from django.http import JsonResponse, HttpResponse
from datetime import datetime, timedelta
from django.utils import timezone
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Import your models
from core.models import (
    Organization, OrganizationType, RoleAssignment, OrganizationRole,
    Report
)
from emt.models import EventProposal

@login_required
@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def data_export_filter_view(request):
    """Main data export and filter view with dynamic data type selection"""
    
    # Extract all filters from request
    filters = {
        'data_type': request.GET.get('data_type', ''),
        'status': request.GET.getlist('status'),
        'organization_type': request.GET.get('organization_type', ''),
        'organization': request.GET.get('organization', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'academic_year': request.GET.get('academic_year', ''),
        'search': request.GET.get('search', ''),
        'submitted_by': request.GET.get('submitted_by', ''),
        'role': request.GET.get('role', ''),
        'active_only': request.GET.get('active_only', ''),
        'has_parent': request.GET.get('has_parent', ''),
        'venue': request.GET.get('venue', ''),
        'report_type': request.GET.get('report_type', ''),
        'priority': request.GET.get('priority', ''),
        'needs_finance_approval': request.GET.get('needs_finance_approval', ''),
        'is_big_event': request.GET.get('is_big_event', ''),
        'has_role_assignments': request.GET.get('has_role_assignments', ''),
        'faculty_incharge': request.GET.get('faculty_incharge', ''),
        'event_focus_type': request.GET.get('event_focus_type', ''),
        'created_by_year': request.GET.get('created_by_year', ''),
        'updated_in_days': request.GET.get('updated_in_days', ''),
        
        # Event Proposal specific filters
        'event_title': request.GET.get('event_title', ''),
        'student_coordinators': request.GET.get('student_coordinators', ''),
        'target_audience': request.GET.get('target_audience', ''),
        'committees': request.GET.get('committees', ''),
        'event_date_from': request.GET.get('event_date_from', ''),
        'event_date_to': request.GET.get('event_date_to', ''),
        'fest_fee_min': request.GET.get('fest_fee_min', ''),
        'fest_fee_max': request.GET.get('fest_fee_max', ''),
        
        # Organization specific filters
        'org_name': request.GET.get('org_name', ''),
        'parent_organization': request.GET.get('parent_organization', ''),
        
        # User specific filters
        'username': request.GET.get('username', ''),
        'email': request.GET.get('email', ''),
        'first_name': request.GET.get('first_name', ''),
        'last_name': request.GET.get('last_name', ''),
        'user_role': request.GET.get('user_role', ''),
        
        # Report specific filters
        'report_title': request.GET.get('report_title', ''),
        'report_description': request.GET.get('report_description', ''),
    }

    # Handle export request
    export_format = request.GET.get('export')
    if export_format:
        return handle_data_export(request, filters, export_format)

    # Get filtered data based on data type
    filtered_data = apply_data_type_filters(filters)
    filter_counts = get_comprehensive_filter_counts(filters)
    
    # Prepare context for template
    context = {
        # Data type information
        'data_types': get_data_type_configs(),
        'current_data_type': filters['data_type'],
        
        # Filter options
        'organization_types': OrganizationType.objects.filter(is_active=True).order_by('name'),
        'organizations': get_filtered_organizations(filters.get('organization_type')),
        'academic_years': get_available_academic_years(),
        'available_statuses': get_available_statuses(),
        'available_users': get_available_users(),
        'available_roles': get_available_roles(),
        'available_venues': get_available_venues(),
        'available_report_types': get_available_report_types(),
        'available_event_focus_types': get_available_event_focus_types(),
        'faculty_users': get_faculty_users(),
        
        # Results and counts
        'filtered_data': filtered_data,
        'filter_counts': filter_counts,
        'total_records': len(filtered_data) if hasattr(filtered_data, '__len__') else filtered_data.count(),
        
        # Current filter values (for form persistence)
        'current_filters': filters,
        
        # API URLs
        'api_orgs_url': '/api/organizations-by-type/',
        'api_counts_url': '/api/filter-counts/',
    }

    return render(request, 'core/data_export_filter.html', context)

def get_data_type_configs():
    """Return configuration for each data type including counts and descriptions"""
    return {
        '': {
            'name': 'All Data',
            'icon': 'fa-database',
            'description': 'View all data types with common filters',
            'color': 'primary'
        },
        'event_proposals': {
            'name': 'Event Proposals',
            'icon': 'fa-calendar-check',
            'description': 'Venue, academic year, finance filters',
            'color': 'info'
        },
        'organizations': {
            'name': 'Organizations',
            'icon': 'fa-building',
            'description': 'Type, hierarchy, assignment filters',
            'color': 'warning'
        },
        'users': {
            'name': 'Users',
            'icon': 'fa-users',
            'description': 'Roles, activity, login history',
            'color': 'success'
        },
        'reports': {
            'name': 'Reports',
            'icon': 'fa-file-alt',
            'description': 'Type, status, submission date',
            'color': 'secondary'
        }
    }

def apply_data_type_filters(filters):
    """Apply filters based on selected data type"""
    data_type = filters.get('data_type', '')
    
    if data_type == 'event_proposals':
        return filter_event_proposals(filters)
    elif data_type == 'organizations':
        return filter_organizations(filters)
    elif data_type == 'users':
        return filter_users(filters)
    elif data_type == 'reports':
        return filter_reports(filters)
    else:
        # Return mixed data when no specific type is selected
        return filter_mixed_data(filters)

def filter_event_proposals(filters):
    """Enhanced filtering for event proposals with all specific filters"""
    queryset = EventProposal.objects.select_related(
        'submitted_by', 'organization', 'organization__org_type'
    ).prefetch_related('faculty_incharges').order_by('-created_at')

    # Common filters
    if filters.get('status'):
        queryset = queryset.filter(status__in=filters['status'])
    
    if filters.get('organization_type'):
        queryset = queryset.filter(organization__org_type_id=filters['organization_type'])
    
    if filters.get('organization'):
        queryset = queryset.filter(organization_id=filters['organization'])
    
    if filters.get('submitted_by'):
        queryset = queryset.filter(submitted_by_id=filters['submitted_by'])

    # Date filters
    if filters.get('date_from'):
        try:
            date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__gte=date_from)
        except ValueError:
            pass
    
    if filters.get('date_to'):
        try:
            date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__lte=date_to)
        except ValueError:
            pass

    # Event Proposal specific filters
    if filters.get('event_title'):
        queryset = queryset.filter(event_title__icontains=filters['event_title'])
    
    if filters.get('academic_year'):
        queryset = queryset.filter(academic_year=filters['academic_year'])
    
    if filters.get('venue'):
        queryset = queryset.filter(venue__icontains=filters['venue'])
    
    if filters.get('faculty_incharge'):
        queryset = queryset.filter(faculty_incharges=filters['faculty_incharge'])
    
    if filters.get('event_focus_type'):
        queryset = queryset.filter(event_focus_type__icontains=filters['event_focus_type'])
    
    if filters.get('student_coordinators'):
        queryset = queryset.filter(student_coordinators__icontains=filters['student_coordinators'])
    
    if filters.get('target_audience'):
        queryset = queryset.filter(target_audience__icontains=filters['target_audience'])
    
    if filters.get('committees'):
        queryset = queryset.filter(committees__icontains=filters['committees'])

    # Boolean filters
    if filters.get('needs_finance_approval') == 'true':
        queryset = queryset.filter(needs_finance_approval=True)
    elif filters.get('needs_finance_approval') == 'false':
        queryset = queryset.filter(needs_finance_approval=False)
    
    if filters.get('is_big_event') == 'true':
        queryset = queryset.filter(is_big_event=True)
    elif filters.get('is_big_event') == 'false':
        queryset = queryset.filter(is_big_event=False)

    # Event date range filters
    if filters.get('event_date_from'):
        try:
            event_date_from = datetime.strptime(filters['event_date_from'], '%Y-%m-%d')
            queryset = queryset.filter(event_datetime__date__gte=event_date_from.date())
        except ValueError:
            pass
    
    if filters.get('event_date_to'):
        try:
            event_date_to = datetime.strptime(filters['event_date_to'], '%Y-%m-%d')
            queryset = queryset.filter(event_datetime__date__lte=event_date_to.date())
        except ValueError:
            pass

    # Fee range filters
    if filters.get('fest_fee_min'):
        try:
            queryset = queryset.filter(fest_fee_participants__gte=int(filters['fest_fee_min']))
        except (ValueError, TypeError):
            pass
    
    if filters.get('fest_fee_max'):
        try:
            queryset = queryset.filter(fest_fee_participants__lte=int(filters['fest_fee_max']))
        except (ValueError, TypeError):
            pass

    # Time-based filters
    if filters.get('updated_in_days'):
        try:
            days = int(filters['updated_in_days'])
            cutoff_date = timezone.now() - timedelta(days=days)
            queryset = queryset.filter(updated_at__gte=cutoff_date)
        except (ValueError, TypeError):
            pass

    if filters.get('created_by_year'):
        try:
            year = int(filters['created_by_year'])
            queryset = queryset.filter(created_at__year=year)
        except (ValueError, TypeError):
            pass

    # Search across multiple fields
    if filters.get('search'):
        search_query = filters['search']
        queryset = queryset.filter(
            Q(event_title__icontains=search_query) |
            Q(submitted_by__first_name__icontains=search_query) |
            Q(submitted_by__last_name__icontains=search_query) |
            Q(submitted_by__username__icontains=search_query) |
            Q(submitted_by__email__icontains=search_query) |
            Q(organization__name__icontains=search_query) |
            Q(venue__icontains=search_query) |
            Q(committees__icontains=search_query) |
            Q(target_audience__icontains=search_query) |
            Q(event_focus_type__icontains=search_query) |
            Q(student_coordinators__icontains=search_query)
        )

    return queryset

def filter_organizations(filters):
    """Enhanced filtering for organizations"""
    queryset = Organization.objects.select_related(
        'org_type', 'parent'
    ).prefetch_related('role_assignments').order_by('org_type__name', 'name')

    # Common filters
    if filters.get('organization_type'):
        queryset = queryset.filter(org_type_id=filters['organization_type'])

    # Organization specific filters
    if filters.get('org_name'):
        queryset = queryset.filter(name__icontains=filters['org_name'])
    
    if filters.get('parent_organization'):
        queryset = queryset.filter(parent_id=filters['parent_organization'])

    # Boolean filters
    if filters.get('active_only') == 'true':
        queryset = queryset.filter(is_active=True)
    elif filters.get('active_only') == 'false':
        queryset = queryset.filter(is_active=False)
    
    if filters.get('has_parent') == 'true':
        queryset = queryset.filter(parent__isnull=False)
    elif filters.get('has_parent') == 'false':
        queryset = queryset.filter(parent__isnull=True)
    
    if filters.get('has_role_assignments') == 'true':
        queryset = queryset.annotate(num_assignments=Count('role_assignments')).filter(num_assignments__gt=0)
    elif filters.get('has_role_assignments') == 'false':
        queryset = queryset.annotate(num_assignments=Count('role_assignments')).filter(num_assignments=0)

    # Search
    if filters.get('search'):
        search_query = filters['search']
        queryset = queryset.filter(
            Q(name__icontains=search_query) |
            Q(org_type__name__icontains=search_query) |
            Q(parent__name__icontains=search_query)
        )

    return queryset

def filter_users(filters):
    """Enhanced filtering for users"""
    queryset = User.objects.prefetch_related(
        Prefetch('role_assignments', 
                queryset=RoleAssignment.objects.select_related('role', 'organization__org_type'))
    ).order_by('first_name', 'last_name', 'username')

    # Common filters
    if filters.get('organization_type'):
        queryset = queryset.filter(
            role_assignments__organization__org_type_id=filters['organization_type']
        ).distinct()
    
    if filters.get('organization'):
        queryset = queryset.filter(
            role_assignments__organization_id=filters['organization']
        ).distinct()
    
    if filters.get('role') or filters.get('user_role'):
        role_id = filters.get('role') or filters.get('user_role')
        queryset = queryset.filter(role_assignments__role_id=role_id).distinct()

    # User specific filters
    if filters.get('username'):
        queryset = queryset.filter(username__icontains=filters['username'])
    
    if filters.get('email'):
        queryset = queryset.filter(email__icontains=filters['email'])
    
    if filters.get('first_name'):
        queryset = queryset.filter(first_name__icontains=filters['first_name'])
    
    if filters.get('last_name'):
        queryset = queryset.filter(last_name__icontains=filters['last_name'])

    # Boolean filters
    if filters.get('active_only') == 'true':
        queryset = queryset.filter(is_active=True)
    elif filters.get('active_only') == 'false':
        queryset = queryset.filter(is_active=False)
    
    if filters.get('has_role_assignments') == 'true':
        queryset = queryset.annotate(num_assignments=Count('role_assignments')).filter(num_assignments__gt=0)
    elif filters.get('has_role_assignments') == 'false':
        queryset = queryset.annotate(num_assignments=Count('role_assignments')).filter(num_assignments=0)

    # Date filters
    if filters.get('date_from'):
        try:
            date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
            queryset = queryset.filter(date_joined__date__gte=date_from)
        except ValueError:
            pass
    
    if filters.get('date_to'):
        try:
            date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
            queryset = queryset.filter(date_joined__date__lte=date_to)
        except ValueError:
            pass

    # Time-based filters
    if filters.get('created_by_year'):
        try:
            year = int(filters['created_by_year'])
            queryset = queryset.filter(date_joined__year=year)
        except (ValueError, TypeError):
            pass

    # Search
    if filters.get('search'):
        search_query = filters['search']
        queryset = queryset.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(role_assignments__role__name__icontains=search_query) |
            Q(role_assignments__organization__name__icontains=search_query)
        ).distinct()

    return queryset

def filter_reports(filters):
    """Enhanced filtering for reports"""
    queryset = Report.objects.select_related(
        'submitted_by', 'organization', 'organization__org_type'
    ).order_by('-created_at')

    # Common filters
    if filters.get('status'):
        queryset = queryset.filter(status__in=filters['status'])
    
    if filters.get('organization_type'):
        queryset = queryset.filter(organization__org_type_id=filters['organization_type'])
    
    if filters.get('organization'):
        queryset = queryset.filter(organization_id=filters['organization'])
    
    if filters.get('submitted_by'):
        queryset = queryset.filter(submitted_by_id=filters['submitted_by'])

    # Report specific filters
    if filters.get('report_title'):
        queryset = queryset.filter(title__icontains=filters['report_title'])
    
    if filters.get('report_description'):
        queryset = queryset.filter(description__icontains=filters['report_description'])
    
    if filters.get('report_type'):
        queryset = queryset.filter(report_type=filters['report_type'])

    # Date filters
    if filters.get('date_from'):
        try:
            date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__gte=date_from)
        except ValueError:
            pass
    
    if filters.get('date_to'):
        try:
            date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__lte=date_to)
        except ValueError:
            pass

    # Time-based filters
    if filters.get('updated_in_days'):
        try:
            days = int(filters['updated_in_days'])
            cutoff_date = timezone.now() - timedelta(days=days)
            queryset = queryset.filter(updated_at__gte=cutoff_date)
        except (ValueError, TypeError):
            pass

    if filters.get('created_by_year'):
        try:
            year = int(filters['created_by_year'])
            queryset = queryset.filter(created_at__year=year)
        except (ValueError, TypeError):
            pass

    # Search
    if filters.get('search'):
        search_query = filters['search']
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(submitted_by__first_name__icontains=search_query) |
            Q(submitted_by__last_name__icontains=search_query) |
            Q(submitted_by__username__icontains=search_query) |
            Q(submitted_by__email__icontains=search_query) |
            Q(organization__name__icontains=search_query)
        )

    return queryset

def filter_mixed_data(filters):
    """Filter mixed data when no specific data type is selected"""
    mixed_data = []
    limit_per_type = 25

    # Get limited data from each type
    event_proposals = filter_event_proposals(filters)[:limit_per_type]
    organizations = filter_organizations(filters)[:limit_per_type]
    users = filter_users(filters)[:limit_per_type]
    reports = filter_reports(filters)[:limit_per_type]

    # Format event proposals
    for proposal in event_proposals:
        mixed_data.append({
            'data_type': 'event_proposal',
            'display_name': proposal.event_title or 'Untitled Event',
            'display_status': proposal.status,
            'display_status_text': proposal.get_status_display() if hasattr(proposal, 'get_status_display') else proposal.status.title(),
            'display_organization': proposal.organization.name if proposal.organization else 'N/A',
            'display_date': proposal.created_at.strftime('%b %d, %Y'),
            'display_user': proposal.submitted_by.get_full_name() if proposal.submitted_by else 'N/A',
            'original_object': proposal
        })

    # Format organizations
    for org in organizations:
        mixed_data.append({
            'data_type': 'organization',
            'display_name': org.name,
            'display_status': 'active' if org.is_active else 'inactive',
            'display_status_text': 'Active' if org.is_active else 'Inactive',
            'display_organization': org.org_type.name if org.org_type else 'N/A',
            'display_date': getattr(org, 'created_at', timezone.now()).strftime('%b %d, %Y'),
            'display_user': 'System',
            'original_object': org
        })

    # Format users
    for user in users:
        roles_list = [ra.role.name for ra in user.role_assignments.all()[:3]]
        mixed_data.append({
            'data_type': 'user',
            'display_name': user.get_full_name() or user.username,
            'display_status': 'active' if user.is_active else 'inactive',
            'display_status_text': 'Active' if user.is_active else 'Inactive',
            'display_organization': ', '.join(roles_list) if roles_list else 'No roles',
            'display_date': user.date_joined.strftime('%b %d, %Y'),
            'display_user': user.email or 'N/A',
            'original_object': user
        })

    # Format reports
    for report in reports:
        mixed_data.append({
            'data_type': 'report',
            'display_name': report.title,
            'display_status': report.status,
            'display_status_text': report.get_status_display() if hasattr(report, 'get_status_display') else report.status.title(),
            'display_organization': report.organization.name if report.organization else 'N/A',
            'display_date': report.created_at.strftime('%b %d, %Y'),
            'display_user': report.submitted_by.get_full_name() if report.submitted_by else 'System',
            'original_object': report
        })

    # Sort mixed data by priority and date
    def sort_key(item):
        type_priority = {'event_proposal': 1, 'report': 2, 'user': 3, 'organization': 4}
        priority = type_priority.get(item['data_type'], 5)
        date_attr = getattr(item['original_object'], 'created_at', getattr(item['original_object'], 'date_joined', None))
        date_score = date_attr.timestamp() if date_attr else 0
        return (priority, -date_score)

    mixed_data.sort(key=sort_key)
    return mixed_data

def get_comprehensive_filter_counts(filters):
    """Get counts for each data type with current filters"""
    counts = {}
    
    # Create a copy of filters without data_type to get counts for all types
    count_filters = {k: v for k, v in filters.items() if k != 'data_type'}
    
    try:
        counts['event_proposals'] = filter_event_proposals(count_filters).count()
    except Exception as e:
        print(f"Error counting event proposals: {e}")
        counts['event_proposals'] = 0
    
    try:
        counts['organizations'] = filter_organizations(count_filters).count()
    except Exception as e:
        print(f"Error counting organizations: {e}")
        counts['organizations'] = 0
    
    try:
        counts['users'] = filter_users(count_filters).count()
    except Exception as e:
        print(f"Error counting users: {e}")
        counts['users'] = 0
    
    try:
        counts['reports'] = filter_reports(count_filters).count()
    except Exception as e:
        print(f"Error counting reports: {e}")
        counts['reports'] = 0

    counts['total'] = sum(counts.values())
    return counts

# Keep all your existing helper functions
def get_filtered_organizations(org_type_id=None):
    queryset = Organization.objects.filter(is_active=True).select_related('org_type', 'parent')
    if org_type_id:
        try:
            queryset = queryset.filter(org_type_id=int(org_type_id))
        except (ValueError, TypeError):
            pass
    return queryset.order_by('org_type__name', 'name')

def get_available_users():
    return User.objects.filter(is_active=True).prefetch_related(
        'role_assignments__role',
        'role_assignments__organization'
    ).order_by('first_name', 'last_name', 'username')

def get_faculty_users():
    return User.objects.filter(
        Q(role_assignments__role__name__icontains='faculty') |
        Q(groups__name__icontains='faculty')
    ).distinct().order_by('first_name', 'last_name', 'username')

def get_available_roles():
    return OrganizationRole.objects.filter(is_active=True).order_by('name')

def get_available_venues():
    venues = EventProposal.objects.exclude(
        venue__isnull=True
    ).exclude(
        venue__exact=''
    ).values_list('venue', flat=True).distinct()
    return sorted(set(filter(None, venues)))

def get_available_event_focus_types():
    focus_types = EventProposal.objects.exclude(
        event_focus_type__isnull=True
    ).exclude(
        event_focus_type__exact=''
    ).values_list('event_focus_type', flat=True).distinct()
    return sorted(set(filter(None, focus_types)))

def get_available_report_types():
    try:
        return Report.REPORT_TYPE_CHOICES
    except AttributeError:
        report_types = Report.objects.exclude(
            report_type__isnull=True
        ).exclude(
            report_type__exact=''
        ).values_list('report_type', flat=True).distinct()
        return [(rt, rt.replace('_', ' ').title()) for rt in report_types if rt]

def get_available_academic_years():
    years = set()
    emt_years = EventProposal.objects.exclude(
        academic_year__isnull=True
    ).exclude(
        academic_year__exact=''
    ).values_list('academic_year', flat=True).distinct()
    years.update(emt_years)

    if not years:
        current_year = timezone.now().year
        for i in range(10):
            start_year = current_year - i
            years.add(f"{start_year}-{start_year + 1}")

    return sorted(years, reverse=True)

def get_available_statuses():
    statuses = set()
    
    # Get EventProposal statuses
    try:
        if hasattr(EventProposal, 'Status'):
            statuses.update(EventProposal.Status.choices)
        else:
            emt_statuses = EventProposal.objects.exclude(
                status__isnull=True
            ).values_list('status', flat=True).distinct()
            statuses.update([(s, s.replace('_', ' ').title()) for s in emt_statuses])
    except AttributeError:
        pass

    # Get Report statuses
    try:
        if hasattr(Report, 'STATUS_CHOICES'):
            statuses.update(Report.STATUS_CHOICES)
    except AttributeError:
        pass

    unique_statuses = list(set(statuses))
    return sorted(unique_statuses, key=lambda x: x[1] if isinstance(x, tuple) else str(x))

def handle_data_export(request, filters, export_format):
    """Handle data export in various formats"""
    data_type = filters.get('data_type', 'mixed')
    filtered_data = apply_data_type_filters(filters)

    # Generate filename
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename_parts = [data_type.replace('_', '-') if data_type else 'all-data']

    if filters.get('organization'):
        try:
            org = Organization.objects.get(id=filters['organization'])
            filename_parts.append(org.name.replace(' ', '-').lower())
        except:
            pass

    if filters.get('academic_year'):
        filename_parts.append(filters['academic_year'].replace('-', ''))

    filename_parts.append(timestamp)
    filename = '_'.join(filename_parts)

    if export_format == 'csv':
        return export_to_csv(filtered_data, data_type, filename, filters)
    elif export_format == 'excel':
        return export_to_excel(filtered_data, data_type, filename, filters)
    elif export_format == 'pdf':
        return export_to_pdf(filtered_data, data_type, filename, filters)
    else:
        return JsonResponse({'error': 'Invalid export format'}, status=400)

def export_to_csv(data, data_type, filename, filters):
    """Export data to CSV format"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    # Add BOM for proper Excel UTF-8 support
    response.write('\ufeff')
    writer = csv.writer(response)

    # Write metadata
    writer.writerow([f'# Export Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([f'# Data Type: {data_type.replace("_", " ").title()}'])
    writer.writerow([f'# Total Records: {len(data) if hasattr(data, "__len__") else "Unknown"}'])
    writer.writerow([])

    if data_type == 'event_proposals':
        headers = [
            'ID', 'Title', 'Organization', 'Organization Type', 'Status', 'Submitted By',
            'Submitted By Email', 'Date Submitted', 'Last Updated', 'Academic Year',
            'Venue', 'Event Focus Type', 'Target Audience', 'Faculty In-Charges',
            'Student Coordinators', 'Needs Finance Approval', 'Is Big Event',
            'Committees', 'Event Date/Time', 'Fest Fee Participants'
        ]
        writer.writerow(headers)

        for item in data:
            faculty_names = ', '.join([f.get_full_name() or f.username for f in item.faculty_incharges.all()])
            writer.writerow([
                item.id,
                item.event_title or 'N/A',
                item.organization.name if item.organization else 'N/A',
                item.organization.org_type.name if item.organization and item.organization.org_type else 'N/A',
                item.get_status_display() if hasattr(item, 'get_status_display') else item.status,
                item.submitted_by.get_full_name() if item.submitted_by else 'N/A',
                item.submitted_by.email if item.submitted_by else 'N/A',
                item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else 'N/A',
                item.updated_at.strftime('%Y-%m-%d %H:%M') if item.updated_at else 'N/A',
                item.academic_year or 'N/A',
                item.venue or 'N/A',
                item.event_focus_type or 'N/A',
                item.target_audience or 'N/A',
                faculty_names or 'N/A',
                item.student_coordinators or 'N/A',
                'Yes' if getattr(item, 'needs_finance_approval', False) else 'No',
                'Yes' if getattr(item, 'is_big_event', False) else 'No',
                getattr(item, 'committees', '') or 'N/A',
                item.event_datetime.strftime('%Y-%m-%d %H:%M') if hasattr(item, 'event_datetime') and item.event_datetime else 'N/A',
                getattr(item, 'fest_fee_participants', '') or 'N/A'
            ])

    elif data_type == 'organizations':
        headers = [
            'ID', 'Name', 'Organization Type', 'Parent Organization', 'Is Active',
            'Created At', 'Role Assignments Count'
        ]
        writer.writerow(headers)

        for item in data:
            role_count = item.role_assignments.count() if hasattr(item, 'role_assignments') else 0
            writer.writerow([
                item.id,
                item.name,
                item.org_type.name if item.org_type else 'N/A',
                item.parent.name if item.parent else 'N/A',
                'Yes' if item.is_active else 'No',
                getattr(item, 'created_at', timezone.now()).strftime('%Y-%m-%d %H:%M'),
                role_count
            ])

    elif data_type == 'users':
        headers = [
            'ID', 'Username', 'First Name', 'Last Name', 'Email', 'Is Active',
            'Date Joined', 'Last Login', 'Roles', 'Organizations'
        ]
        writer.writerow(headers)

        for item in data:
            roles = [ra.role.name for ra in item.role_assignments.all()]
            organizations = [ra.organization.name for ra in item.role_assignments.all()]
            writer.writerow([
                item.id,
                item.username,
                item.first_name or 'N/A',
                item.last_name or 'N/A',
                item.email or 'N/A',
                'Yes' if item.is_active else 'No',
                item.date_joined.strftime('%Y-%m-%d %H:%M'),
                item.last_login.strftime('%Y-%m-%d %H:%M') if item.last_login else 'Never',
                ', '.join(roles) or 'No roles',
                ', '.join(organizations) or 'No organizations'
            ])

    elif data_type == 'reports':
        headers = [
            'ID', 'Title', 'Description', 'Organization', 'Status', 'Report Type',
            'Submitted By', 'Created At', 'Updated At'
        ]
        writer.writerow(headers)

        for item in data:
            writer.writerow([
                item.id,
                item.title,
                (item.description[:100] + '...' if len(item.description) > 100 else item.description) if hasattr(item, 'description') and item.description else 'N/A',
                item.organization.name if item.organization else 'N/A',
                item.get_status_display() if hasattr(item, 'get_status_display') else getattr(item, 'status', 'N/A'),
                getattr(item, 'report_type', 'N/A') or 'N/A',
                item.submitted_by.get_full_name() if item.submitted_by else 'N/A',
                item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else 'N/A',
                item.updated_at.strftime('%Y-%m-%d %H:%M') if hasattr(item, 'updated_at') and item.updated_at else 'N/A'
            ])

    else:  # Mixed data
        headers = [
            'Data Type', 'Name', 'Status', 'Organization', 'Date', 'User'
        ]
        writer.writerow(headers)

        for item in data:
            writer.writerow([
                item['data_type'].replace('_', ' ').title(),
                item['display_name'],
                item['display_status_text'],
                item['display_organization'],
                item['display_date'],
                item['display_user']
            ])

    return response

def export_to_excel(data, data_type, filename, filters):
    """Export data to Excel format with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = data_type.replace('_', ' ').title()[:31]  # Excel sheet name limit

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Write metadata
    ws.append([f"Export Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"Data Type: {data_type.replace('_', ' ').title()}"])
    ws.append([f"Total Records: {len(data) if hasattr(data, '__len__') else 'Unknown'}"])
    ws.append([])

    start_row = 5

    if data_type == 'event_proposals':
        headers = [
            'ID', 'Title', 'Organization', 'Organization Type', 'Status', 'Submitted By',
            'Submitted By Email', 'Date Submitted', 'Last Updated', 'Academic Year',
            'Venue', 'Event Focus Type', 'Target Audience', 'Faculty In-Charges',
            'Student Coordinators', 'Needs Finance Approval', 'Is Big Event',
            'Committees', 'Event Date/Time', 'Fest Fee Participants'
        ]
        ws.append(headers)

        # Apply header formatting
        for cell in ws[start_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_num, item in enumerate(data, start=start_row + 1):
            faculty_names = ', '.join([f.get_full_name() or f.username for f in item.faculty_incharges.all()])
            row_data = [
                item.id,
                item.event_title or 'N/A',
                item.organization.name if item.organization else 'N/A',
                item.organization.org_type.name if item.organization and item.organization.org_type else 'N/A',
                item.get_status_display() if hasattr(item, 'get_status_display') else item.status,
                item.submitted_by.get_full_name() if item.submitted_by else 'N/A',
                item.submitted_by.email if item.submitted_by else 'N/A',
                item.created_at if item.created_at else 'N/A',
                item.updated_at if item.updated_at else 'N/A',
                item.academic_year or 'N/A',
                item.venue or 'N/A',
                item.event_focus_type or 'N/A',
                item.target_audience or 'N/A',
                faculty_names or 'N/A',
                item.student_coordinators or 'N/A',
                'Yes' if getattr(item, 'needs_finance_approval', False) else 'No',
                'Yes' if getattr(item, 'is_big_event', False) else 'No',
                getattr(item, 'committees', '') or 'N/A',
                item.event_datetime if hasattr(item, 'event_datetime') and item.event_datetime else 'N/A',
                getattr(item, 'fest_fee_participants', '') or 'N/A'
            ]
            ws.append(row_data)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response

def export_to_pdf(data, data_type, filename, filters):
    """Export data to PDF format"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1  # Center alignment
    )
    
    elements.append(Paragraph(f"Data Export Report - {data_type.replace('_', ' ').title()}", title_style))
    elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    if data_type == 'event_proposals':
        table_data = [['Title', 'Organization', 'Status', 'Submitted By', 'Date', 'Academic Year', 'Venue']]

        for item in data:
            table_data.append([
                (item.event_title or 'N/A')[:30] + ('...' if len(item.event_title or '') > 30 else ''),
                (item.organization.name if item.organization else 'N/A')[:20] + ('...' if len(item.organization.name if item.organization else '') > 20 else ''),
                item.get_status_display() if hasattr(item, 'get_status_display') else item.status,
                (item.submitted_by.get_full_name() if item.submitted_by else 'N/A')[:20],
                item.created_at.strftime('%Y-%m-%d') if item.created_at else 'N/A',
                item.academic_year or 'N/A',
                (item.venue or 'N/A')[:20] + ('...' if len(item.venue or '') > 20 else '')
            ])

    elif data_type == 'organizations':
        table_data = [['Name', 'Type', 'Parent', 'Active', 'Role Assignments']]

        for item in data:
            role_count = item.role_assignments.count() if hasattr(item, 'role_assignments') else 0
            table_data.append([
                item.name[:30] + ('...' if len(item.name) > 30 else ''),
                item.org_type.name if item.org_type else 'N/A',
                item.parent.name[:20] if item.parent else 'N/A',
                'Yes' if item.is_active else 'No',
                str(role_count)
            ])

    elif data_type == 'users':
        table_data = [['Name', 'Username', 'Email', 'Active', 'Date Joined', 'Roles']]

        for item in data:
            roles = [ra.role.name for ra in item.role_assignments.all()[:2]]  # Limit to 2 roles for PDF
            table_data.append([
                (item.get_full_name() or item.username)[:25],
                item.username[:20],
                item.email[:25] if item.email else 'N/A',
                'Yes' if item.is_active else 'No',
                item.date_joined.strftime('%Y-%m-%d'),
                ', '.join(roles)[:30] if roles else 'No roles'
            ])

    else:  # Mixed or other data types
        table_data = [['Type', 'Name', 'Status', 'Organization', 'Date']]

        for item in data:
            table_data.append([
                item['data_type'].replace('_', ' ').title(),
                item['display_name'][:30],
                item['display_status_text'],
                item['display_organization'][:25],
                item['display_date']
            ])

    # Create and style the table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

# API endpoints for AJAX requests
@login_required
def api_organizations_by_type(request):
    """API endpoint to get organizations filtered by type"""
    org_type_id = request.GET.get('org_type', '')
    queryset = Organization.objects.filter(is_active=True).select_related('org_type', 'parent')

    if org_type_id:
        try:
            queryset = queryset.filter(org_type_id=int(org_type_id))
        except (ValueError, TypeError):
            pass

    organizations = []
    for org in queryset.order_by('name')[:100]:  # Limit results for performance
        organizations.append({
            'id': org.id,
            'name': org.name,
            'org_type__name': org.org_type.name if org.org_type else 'N/A',
        })

    return JsonResponse(organizations, safe=False)

@login_required
def api_filter_counts(request):
    """API endpoint to get real-time filter counts"""
    try:
        filters = {}
        for key in request.GET:
            if key == 'status':
                filters[key] = request.GET.getlist(key)
            else:
                filters[key] = request.GET.get(key)

        counts = get_comprehensive_filter_counts(filters)
        return JsonResponse({
            **counts,
            'last_updated': timezone.now().isoformat(),
            'success': True
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'success': False
        }, status=500)

@login_required
def api_users_by_role(request):
    """API endpoint to get users filtered by role"""
    role_id = request.GET.get('role', '')
    queryset = User.objects.filter(is_active=True).prefetch_related('role_assignments__role')

    if role_id:
        try:
            queryset = queryset.filter(role_assignments__role_id=int(role_id)).distinct()
        except (ValueError, TypeError):
            pass

    users = []
    for user in queryset.order_by('first_name', 'last_name', 'username')[:100]:
        users.append({
            'id': user.id,
            'name': user.get_full_name() or user.username,
            'email': user.email or '',
        })

    return JsonResponse(users, safe=False)

@login_required
def api_search_suggestions(request):
    """API endpoint for search suggestions based on data type"""
    data_type = request.GET.get('data_type', '')
    query = request.GET.get('q', '').strip()
    
    suggestions = []
    
    if len(query) >= 2:  # Only search if query is at least 2 characters
        if data_type == 'event_proposals':
            # Search in event titles
            events = EventProposal.objects.filter(
                event_title__icontains=query
            ).values_list('event_title', flat=True)[:10]
            suggestions.extend([{'type': 'event_title', 'value': title} for title in events if title])
            
        elif data_type == 'organizations':
            # Search in organization names
            orgs = Organization.objects.filter(
                name__icontains=query,
                is_active=True
            ).values_list('name', flat=True)[:10]
            suggestions.extend([{'type': 'organization', 'value': name} for name in orgs])
            
        elif data_type == 'users':
            # Search in user names and usernames
            users = User.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(username__icontains=query),
                is_active=True
            ).values('first_name', 'last_name', 'username')[:10]
            
            for user in users:
                full_name = f"{user['first_name']} {user['last_name']}".strip()
                suggestions.append({
                    'type': 'user', 
                    'value': full_name if full_name else user['username']
                })
    
    return JsonResponse({'suggestions': suggestions})
# ---------------------------------------------
#           Switch View (Admin)
# ---------------------------------------------

def is_admin(user):
    """Check if user is admin or superuser"""
    return user.is_staff or user.is_superuser

@login_required
@user_passes_test(is_admin)
def switch_user_view(request):
    """Display the switch user interface"""
    users = User.objects.filter(is_active=True).select_related().order_by('username')
    
    # Get current impersonated user if any
    impersonated_user = None
    if 'impersonate_user_id' in request.session:
        try:
            impersonated_user = User.objects.get(id=request.session['impersonate_user_id'])
        except User.DoesNotExist:
            # Clean up invalid session data
            del request.session['impersonate_user_id']
    
    context = {
        'users': users,
        'impersonated_user': impersonated_user,
        'original_user': request.user,
    }
    return render(request, 'core/switch_user.html', context)

@login_required
@user_passes_test(is_admin)
@require_POST
def impersonate_user(request):
    """Start impersonating a user"""
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID is required'})
        
        target_user = get_object_or_404(User, id=user_id, is_active=True)
        
        # Store the impersonation in session
        request.session['impersonate_user_id'] = target_user.id
        request.session['original_user_id'] = request.user.id
        
        messages.success(request, f'Now viewing as: {target_user.get_full_name() or target_user.username}')
        
        return JsonResponse({
            'success': True,
            'message': f'Now impersonating {target_user.get_full_name() or target_user.username}',
            'redirect_url': '/core-admin/'  # Adjust to your dashboard URL
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def stop_impersonation(request):
    """Stop impersonating and return to original user"""
    if 'impersonate_user_id' in request.session:
        del request.session['impersonate_user_id']
        if 'original_user_id' in request.session:
            del request.session['original_user_id']
        messages.success(request, 'Stopped impersonation')
    
    return redirect('core-admin')  # Adjust to your dashboard URL

@login_required
@user_passes_test(is_admin)
@require_POST
def search_users_api(request):
    """API endpoint for quick user search"""
    try:
        data = json.loads(request.body)
        search_term = data.get('search', '').strip()
        
        if len(search_term) < 2:
            return JsonResponse({'users': []})
        
        # Search users by name, username, or email
        users = User.objects.filter(
            Q(username__icontains=search_term) |
            Q(first_name__icontains=search_term) |
            Q(last_name__icontains=search_term) |
            Q(email__icontains=search_term),
            is_active=True
        ).exclude(
            id=request.user.id  # Don't include current user
        ).order_by('first_name', 'last_name', 'username')[:10]  # Limit to 10 results
        
        users_data = []
        for user in users:
            users_data.append({
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'email': user.email,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser
            })
        
        return JsonResponse({'users': users_data})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(is_admin)
def get_recent_users_api(request):
    """Get recently switched users"""
    try:
        # Get recent impersonation history (you might want to store this in a model)
        # For now, returning some sample data
        recent_users = [
            {
                'id': 1,
                'name': 'John Doe',
                'email': 'john@example.com',
                'last_switch': '2024-01-15'
            }
        ]
        
        return JsonResponse({'recent_users': recent_users})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Enhanced impersonate_user view with recent users tracking
@login_required
@user_passes_test(is_admin)
@require_POST
def impersonate_user(request):
    """Start impersonating a user (enhanced with recent users tracking)"""
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID is required'})
        
        target_user = get_object_or_404(User, id=user_id, is_active=True)
        
        # Optional: Prevent impersonating superusers (uncomment if needed)
        # if target_user.is_superuser and not request.user.is_superuser:
        #     return JsonResponse({'success': False, 'error': 'Cannot impersonate superusers'})
        
        # Store the impersonation in session
        request.session['impersonate_user_id'] = target_user.id
        request.session['original_user_id'] = request.user.id
        
        # Store in recent users (in localStorage via JavaScript, or you could use a model)
        messages.success(request, f'Now viewing as: {target_user.get_full_name() or target_user.username}')
        
        return JsonResponse({
            'success': True,
            'message': f'Now impersonating {target_user.get_full_name() or target_user.username}',
            'user': {
                'id': target_user.id,
                'username': target_user.username,
                'full_name': target_user.get_full_name(),
                'email': target_user.email
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})